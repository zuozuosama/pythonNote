import openpyxl
import pymysql

def addToDB(worksheet):
    column = worksheet.max_column
    if column == 16:
        time = 2000
    else:
        time = 2014
    a=time
    # 按行读取数据 list(sh.rows)
    for cases in list(worksheet.rows)[1:]:
        area_code = cases[0].value
        area = cases[1].value
        for x in range(2, column, 1):
            consume = cases[x].value
            cls = 'insert into milk_consume (area_code,area,time,consume)value ("%s","%s","%s","%s")' % (
                str(area_code), str(area), str(time), str(consume))
            cursor.execute(cls)
            time += 1
        time = a
    return

def addToDBGeneral(worksheet,col):
    column = worksheet.max_column
    time=2000
    for cases in list(worksheet.rows)[1:]:
        country_name = cases[0].value
        for x in range(1, column, 1):
            data = cases[x].value
            try:
                cls = 'insert into %s (country_name,time,data)value ("%s","%s","%s")' % (
                    col,str(country_name), str(time), str(data))
                cursor.execute(cls)
            except Exception as e:
                print(e)
                print(len(country_name))
                print(cls)
                return
            time += 1
        time = 2000
    return

def selectAndWriteExcel():
    # 创建一个工作簿
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet('乳品消费量2000-2017')
    sql = 'SELECT * FROM milk_consume WHERE milk_consume.area in (SELECT a.area FROM (SELECT area,count(1)as count FROM milk_consume GROUP BY area HAVING(count)=18) a) ORDER BY milk_consume.area,time;'
    cursor.execute(sql)
    info = cursor.fetchall()
    data_head =['area','time','consume']
    length = len(data_head)
    for i,x in enumerate(data_head):
        sheet.cell(row=1,column=i+1,value=x)

    for i,row in enumerate(info):
        for x in range(length):
            sheet.cell(row=i+2,column=x+1,value=row[x+2])

    wb.save('人均年乳品消费量.xlsx')

if __name__ == '__main__':
    wb = openpyxl.load_workbook('C:\\Users\\lizongqi\\Desktop\\论文数据.xlsx')
    # 第二步：选取表单
    sh = wb['乳品消费量2000-2013']
    sh1 = wb['乳品消费量2014-2017']
    # 链接sql
    db = pymysql.connect(host='localhost', port=3306, user='root', passwd='root', db='syndata', charset='utf8')
    cursor = db.cursor()

    wsName = ['人均GDP','城镇化水平','粗出生率','65和65岁以上人口比重']
    cols = ['avg_gdp','urbanization_level','birth_rate','aging_degree']
    for i,x in enumerate(wsName):
        ws = wb[x]
        addToDBGeneral(ws,cols[i])
    # addToDB(sh)
    # addToDB(sh1)
    # selectAndWriteExcel()

    db.commit()
    db.close()
    # 关闭工作薄
    wb.close()
